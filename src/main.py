from constants import SHEET_NAME, NAME_CELL, MONTH_CELL, START_ROW, DAY_COLUMN, DAY_OF_WEEK_COLUMN
from messages import output_message, Message
from validator import Validator

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

import argparse
import datetime
import os


def read_args():
    parser = argparse.ArgumentParser(description='translator')

    parser.add_argument('file_path', help='target excel file')

    now = datetime.datetime.now()

    default_month = now.month
    month_choices = [i for i in range(1, 13)]
    parser.add_argument('-m', '--month', type=int, default=default_month,
        choices=month_choices, help='target month')

    default_year = now.year
    year_choices = [i for i in range(1970, now.year + 2)]
    parser.add_argument('-y', '--year', type=int, default=default_year,
        choices=year_choices, help='target year')

    args = parser.parse_args()
    file_path = os.path.join(os.getcwd(), args.file_path)
    month = args.month
    year = args.year

    return file_path, month, year


def main():
    file_path, month, year = read_args()

    if not os.path.exists(file_path):
        output_message(Message.FILE_NO_EXISTS_ERROR, file_path)
        return

    file_name, file_ext = os.path.splitext(os.path.basename(file_path))

    if file_ext != '.xlsx':
        output_message(Message.EXTENSION_ERROR, file_path)
        return

    validator = Validator(year, month)
    validator.validate_file_name(file_name)

    workbook = load_workbook(file_path)
    sheet = workbook[SHEET_NAME]

    dt = from_excel(sheet.cell(row=MONTH_CELL[0], column=MONTH_CELL[1]).value)
    validator.validate_month(dt)

    name = sheet.cell(row=NAME_CELL[0], column=NAME_CELL[1]).value
    validator.validate_name(name)

    last_row = START_ROW + validator.last_day - 1
    for row in range(START_ROW, last_row + 1):
        dt = sheet.cell(row=row, column=DAY_COLUMN).value
        day_of_week = sheet.cell(row=row, column=DAY_OF_WEEK_COLUMN).value
        validator.validate_day_of_week(dt.day, day_of_week, (row, DAY_OF_WEEK_COLUMN))


if __name__ == '__main__':
    main()
